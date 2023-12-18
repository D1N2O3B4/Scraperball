import pandas as pd
from pandas import DataFrame
# import msoffcrypto
import openpyxl as xl
from openpyxl.styles import PatternFill, Alignment, Font


import datetime
import os


from generator_utils import copy_range, get_league_match
from utils import resource_path, to_2_dec

colors = {
    "red": "FF0000",
    "light-blue": "56B0F0",
    "orange": "FF8000",
    "green": "059C00",
    "light-green": "C6EFCE",
    "light-gray": "A6A6A6",
    "deep-blue": "000099",
    "deep-gray": "808170",
    "black-gray": "424A41",
    "brown": "663300",
    "black": "000000",
    "pink": "FF00FF",
    "blue": "356DB9",
    "yellow": "FFFF00",
    
}

def generate(df : DataFrame):
    # df = pd.read_csv('./output.csv')
    try:
        df = df.drop(columns=['Unnamed: 0'], axis=1, ) 
    except Exception as e:
        # print(e)
        pass
    
    df.sort_values(by='League', inplace=True)
    # print(df)
    
    leagues = {
        
    }
    
    rows = df.iterrows()
    
    for row in rows:
        row = row[1]
        home = row['Home']
        away = row['Away']
        
        league_match = get_league_match()
        
        for key in row.keys():
            if key == 'League': continue
            if key == 'Home' or key == 'Away':
                league_match['teams'] = f'{row["Home"]} vs {row["Away"]}'
                continue
            # if key == "Hand2" or key == "Diff2" or key == "TG-HT" or key == "TG-2H": continue
            league_match[key] = row[key]
        if row['League'] not in leagues:
            leagues[f'{row["League"]}'] = []
    
        leagues[f'{row["League"]}'].append(league_match)
    
    wb = xl.load_workbook(resource_path('Myfile-decrypted.xlsx'))
    ws = wb.active
    
    headers = ws['A2' : 'BI2'][0]
    headers = [x.value for x in headers]
    current = 2
    start = None
    for key in leagues.keys():

        # if key == "Hand2" or key == "Diff2" or key == "TG-HT" or key == "TG-2H": continue
        
        if start is not None:
            ws.row_dimensions.group(start, current - 1, hidden=False)

        copy_range('A2:BI2', ws, current - 2)
        for i, header in enumerate(headers):
            if i == 0:
                if key == "Diff2" or key == "Hand2":
                    key = key.replace('2', '')
                ws.cell(row=current, column=i+1).value = key
                continue
            # ws.cell(row=current, column=(i+1)).value = header
        current += 1
        start = current
        for match in leagues[key]:
            # ws.cell(row=current, culumn=)
            for prop in match.keys():
                if prop == 'teams':
                    ws.cell(row=current, column=1).value = match['teams']
                    continue
                

                col = headers.index(prop) + 1
                
                cell = ws.cell(row=current, column=col)

                val = match[prop]
                cell.alignment = Alignment(horizontal='center', vertical="center")

                if val == '' and not (prop == "Stat" or prop == "For" or prop == "Bet"):
                    cell.value = val
                    continue
                
                if prop == "Res":
                    if match[prop] == 'SHW' or match[prop] == "MHW":
                        cell.fill = PatternFill("solid", fgColor=colors['light-blue'])
                    elif match[prop] == 'SAW' or match[prop] == "MAW":
                        cell.fill = PatternFill("solid", fgColor=colors['red'])
                    else:
                        cell.fill = PatternFill("solid", fgColor=colors['light-gray'])
                    continue

                if prop == "H" or prop == "A":
                    if val == 'W':
                        cell.fill = PatternFill("solid", fgColor=colors['light-blue'])
                    elif val == 'L':
                        cell.fill = PatternFill("solid", fgColor=colors['red'])
                
                if prop == "H2H":
                    val, score = val
                    if val == '': continue
                    diff = score[0] - score[1]
                    if val >= 1:
                        cell.fill = PatternFill("solid", fgColor=colors['deep-blue'])
                        # if diff >= 2:
                        #     # change font to Myriad Pro with red color font size 8
                        #     cell.font = Font(name='Myriad Pro', size=8, color=colors['yellow'], bold=True)
                            
                        #     pass
                        if diff == 1:
                            cell.font = Font(name='Myriad Pro', size=8, color=colors['red'], bold=True)
                            pass
                    elif val >= 0.33:
                        cell.fill = PatternFill("solid", fgColor=colors['deep-gray'])
                    else:
                        cell.fill = PatternFill("solid", fgColor=colors['red'])
                    cell.value = val
                    continue
                    
                        
                if prop == "5H" or prop == "5A" or prop == "L3H" or prop == "L3A":
                    cell.value = val[0]
                    last_goals = val[1]
                    # print(last_goals)
                    if last_goals >= 7:
                        cell.fill = PatternFill("solid", fgColor=colors['brown'])
                    elif last_goals == 6:
                        cell.fill = PatternFill("solid", fgColor=colors['deep-blue'])
                    elif last_goals == 5:
                        cell.fill = PatternFill("solid", fgColor=colors['black'])
                    elif last_goals == 4:
                        cell.fill = PatternFill("solid", fgColor=colors['pink'])
                    elif last_goals == 3:
                        cell.fill = PatternFill("solid", fgColor=colors['orange'])
                    elif last_goals == 2:
                        cell.fill = PatternFill("solid", fgColor=colors['light-blue'])
                    elif last_goals == 1:
                        cell.fill = PatternFill("solid", fgColor=colors['green'])
                    elif last_goals == 0:
                        cell.fill = PatternFill("solid", fgColor=colors['red'])  
                    continue              
                
                if prop == 'Stat':
                    cell.fill = PatternFill("solid", fgColor=colors['light-green'])
                    continue
                
                if prop == "For":
                    cell.fill = PatternFill("solid", fgColor=colors['light-blue'])
                    continue

                
                if prop == "BF":
                    bf = val
                    bf_val = int((bf[0]/bf[1]) * 100)
                    if bf_val >= 80:
                        cell.fill = PatternFill("solid", fgColor='DA9794')
                    elif bf_val >= 40:
                        cell.fill = PatternFill("solid", fgColor=colors['black-gray'])
                    else:
                        cell.fill = PatternFill("solid", fgColor=colors['blue'])
                    
                    # for_cell = ws.cell(row=current, column=col - 2)
                    # fr = for_cell.value
                    # print(fr)
                    # if int(fr) >= 50:
                    #     bf[0] += 1
                    cell.value = bf_val
                    continue
                
                if prop == "Bet":
                    # if val == "H10" or val == "A10":
                    # if True:
                    # cell.fill = PatternFill("solid", fgColor="0000FF")
                    # cell.fill = PatternFill("solid", fgColor="000000")
                    # print(val, colors["black"])
                    pass
                    
                
                if prop == "Hand":
                    if float(val) == 0.0 or float(val) == -0.0:
                        cell.value = 'L'
                        continue
                    # if float(val) < 0:
                        # cell.fill = PatternFill("solid", fgColor="FF0000")
                    elif float(val) > 0:
                        cell.fill = PatternFill("solid", fgColor=colors['light-blue'])
                    
                if prop == "O-O" or prop == "L-O" or prop == "TGO" or prop == "LO2":

                    if float(val) < 1.95:
                        cell.fill = PatternFill("solid", fgColor=colors['orange'])
                    else:
                        cell.fill = PatternFill("solid", fgColor=colors['red'])
                    
                    # to 2 decimal places
                    cell.value = to_2_dec(val)
                    continue
                
                if prop == "Diff" or prop == "Diff2":
                    if int(val) < 0:
                        cell.fill = PatternFill("solid", fgColor=colors['red'])
                    else:
                        cell.fill = PatternFill("solid", fgColor=colors['light-blue'])
                
                if prop == "TG-HT":
                    # print(val, type(val))
                    val_split = val.split(' ')
                    if len(val_split) == 2: 
                        goals, odds = val_split
                        if float(goals) >= 1:
                            # cell.fill = PatternFill("solid", fgColor=colors['black'])
                            if float(odds) < 1.95:
                                cell.fill = PatternFill("solid", fgColor=colors['orange'])
                            else:
                                cell.fill = PatternFill("solid", fgColor=colors['red'])
                        pass    

                cell.value = match[prop]

            current += 1
    if start is not None:
        ws.row_dimensions.group(start, current - 1, hidden=False)
    
    cell = ws.cell(row=current, column=1)
    cell.value = cell.value
    
    
    try:
        os.mkdir('./output')
        pass
    except OSError as e:
        # print(e)
        pass
    
    now = datetime.datetime.now()
    name = str(now).split('.')[0].replace(':', '')
    # print(name)
    wb.save(f'./output/{name}.xlsx')